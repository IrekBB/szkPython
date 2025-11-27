def main(args):
    skoroszyt = openpyxl.load_workbook(r"C:\Users\irkbe\Downloads\Wyniki_EM_szkoly_lipiec_PP2023_2025.xlsx")
    ws = skoroszyt['SAS']
    print("Max row:", ws.max_row)
    print("Max column:", ws.max_column)
    for i  in range (1,ws.max_row+1):
        if ws.cell(row=i, column=8).value =="LICEUM OGÓLNOKSZTAŁCĄCE NR VII IM. KRZYSZTOFA KAMILA BACZYŃSKIEGO":
            wiersz= i
    for j in range(1, ws.max_column+1):
        if ws.cell(row=2, column=j).value=="informatyka poziom rozszerzony (M)":
            kolumna=j
    print("i,j={},{}".format(wiersz,kolumna))
    for i in range(1,7):
        print (f"{ws.cell(row=3, column=kolumna+i-1).value.strip():<30}", end="")
    print()
    x="0"
    for i in range(6):
            if ws.cell(row=wiersz, column=kolumna+i).value!="":
                print (f"{ws.cell(row=wiersz, column=kolumna+i).value:<30}", end="")
            else: print(f"{x:<30}", end="")
    print()


   



if __name__=="__main__":
    import sys
    import openpyxl
    import numpy as np
    sys.exit(main(sys.argv))